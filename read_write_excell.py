import os, sys
import openpyxl
import locale
from datetime import datetime
from PyQt5.QtWidgets import *
from PyQt5 import QtGui


locale.setlocale(locale.LC_ALL, "ru_RU")
date = datetime.today().strftime("%d.%m.%Y")


class ReadWriteExcel(QWidget):
    def __init__(self, parent=None):
        self.parent = parent
        super(ReadWriteExcel, self).__init__()
        self.make_file = QPushButton('   Почати заповнення "Залишки підрозділи"')
        self.make_file.setIcon(QtGui.QIcon('img/process.png'))
        self.make_file.clicked.connect(self.scrap_write)
        #
        self.make_file_2 = QPushButton('   Почати заповнення "Залишки ООДК"')
        self.make_file_2.setIcon(QtGui.QIcon('img/process.png'))
        self.make_file_2.clicked.connect(self.scrap_write_2)

        main_v_box = QVBoxLayout(self)
        main_v_box.addWidget(self.make_file)
        main_v_box.addWidget(self.make_file_2)

    # for button #1 functions
    def get_dir_path(self):
        self.files_directory_choice = QFileDialog.getExistingDirectory(self, 'Оберіть папку з вихідними файлами')
        return self.files_directory_choice

    def get_file_path(self):
        self.file_main = QFileDialog.getOpenFileName(self, 'Оберіть базовий файл')
        return self.file_main[0]

    def scrap_write(self):
        f_dir = self.get_dir_path()
        m_file = self.get_file_path()
        files = os.listdir(f_dir)
        val_massive = []
        for file in files:
            read_data = openpyxl.load_workbook(f_dir+'/'+file)
            sheet_1 = read_data.sheetnames[0]
            sheet = read_data[sheet_1]
            columns = [2, 3, 4, 5, 10, 12] # ????

            for coll in columns:
                mark = sheet.cell(column=coll, row=6).value
                if mark == None:
                    break
                val_list = []
                count = 0
                for row in range(6, 119):
                    val = sheet.cell(column=coll, row=row).value
                    if val == None:
                        val = 0
                    if type(val) == float or type(val) == int:
                        if count > 22:
                            if val == 0:
                                val = '0,000'
                            else:
                                val = round((val / 1000), 3)
                                val = locale.str(val)
                    count += 1
                    val_list.append(val)
                val_massive.append(val_list)
        #print(val_massive)

        open_file = openpyxl.load_workbook(m_file)
        sheet_main_1 = open_file.sheetnames[0]
        sheet_main = open_file[sheet_main_1]
        columns_main = [x for x in range(4, (len(val_massive)+2)*2) if x % 2 == 0]

        col_index = 0
        for val_l in val_massive:
            column_numb = columns_main[col_index]
            row = 6
            for val in val_l:
                sheet_main.cell(column=column_numb, row=row).value = val
                row += 1
            col_index += 1
        # variant. But ...
        new_filename = m_file.split('.')
        new_filename.insert(1, date)
        new_filename[2] = '.xlsx'
        new_filename = ' '.join(new_filename)
        open_file.save(new_filename)

    # for button #2 functions
    def get_src_file_path(self):
        self.files_directory_choice = QFileDialog.getOpenFileName(self, 'Оберіть вихідний файл')
        return self.files_directory_choice[0]

    def scrap_write_2(self):
        src_file = self.get_src_file_path()
        # src_file = '94 ПРИКЗ Залишки продуктів (по підрозділах) 07.05.2024.xlsx'
        base_file = self.get_file_path()
        # base_file = '94 ПРИКЗ Орган охорони ДК 08.05.2024 – копія.xlsx'
        # read data from source file
        read_data_src_file = openpyxl.load_workbook(src_file)
        sheet_1_src_file = read_data_src_file.sheetnames[0]
        sheet_src_file = read_data_src_file[sheet_1_src_file]
        # columns = [32, 33]
        dict_source_1 = {}
        dict_source_2 = {}

        for row in range(7, 119):
            key = sheet_src_file.cell(column=1, row=row).value
            val_1 = sheet_src_file.cell(column=32, row=row).value
            val_2 = sheet_src_file.cell(column=33, row=row).value
            dict_source_1[key] = val_1
            dict_source_2[key] = val_2

        read_data_base_file = openpyxl.load_workbook(base_file)
        sheet_1_base_file = read_data_base_file.sheetnames[0]
        sheet_base_file = read_data_base_file[sheet_1_base_file]
        ### 1
        for item in dict_source_1.items():
            name, value = item[0], item[1]

            prev_key = None
            for i in range(6, 117):
                key = sheet_base_file.cell(column=1, row=i).value

                if key == None:
                    key = prev_key
                else:
                    prev_key = key

                if key in ['Капуста', 'Морква', 'Буряк', 'Цибуля', 'Часник', 'Огірки, помідори, баклажани']:
                    key = key + ' ' + sheet_base_file.cell(column=2, row=i).value

                if key.lower() == name.lower():
                    row = sheet_base_file.cell(column=1, row=i).row
                    sheet_base_file.cell(column=3, row=row).value = value
                    break
        ### 2
        for item in dict_source_2.items():
            name, value = item[0], item[1]

            prev_key = None
            for i in range(6, 117):
                key = sheet_base_file.cell(column=1, row=i).value

                if key == None:
                    key = prev_key
                else:
                    prev_key = key

                if key in ['Капуста', 'Морква', 'Буряк', 'Цибуля', 'Часник', 'Огірки, помідори, баклажани']:
                    key = key + ' ' + sheet_base_file.cell(column=2, row=i).value

                if key.lower() == name.lower():
                    row = sheet_base_file.cell(column=1, row=i).row
                    sheet_base_file.cell(column=5, row=row).value = value
                    break

        new_filename = base_file.split('.')
        new_filename.insert(1, date)
        new_filename[2] = '.xlsx'
        new_filename = ' '.join(new_filename)
        read_data_base_file.save(new_filename)


class WindowAuthor(QWidget):
    def __init__(self, parent=None):
        self.parent = parent
        super(WindowAuthor, self).__init__()


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__() # initialization widgets and properties Parent class "QDialog"
        self.setWindowTitle('Документ Парсер             (розроблено OlegVostokUA)')
        self.setWindowIcon(QtGui.QIcon('img/software.png'))
        self.resize(600, 300) # set size window
        self.main_widget = QTabWidget()
        self.setCentralWidget(self.main_widget)
        self.main_widget.addTab(ReadWriteExcel(), "Парсер Excel")


if __name__ == '__main__':
    app = QApplication(sys.argv) # create app
    dlgMain = MainWindow() # build object of class "DlgMain" and set here in variable "dlgMain"
    dlgMain.show() # show function
    sys.exit(app.exec_())  # loop app in "sys.exit" func for check logs
