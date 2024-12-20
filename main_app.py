# pip install pyinstaller
# pyinstaller -F -w -i "C:\Users\User\PycharmProjects\ProdPlace\MainApp\main.ico" MainApp.py
import copy
import os
import sys
from datetime import datetime

import openpyxl
import pandas as pd
import locale

from PyQt5.QtGui import QFont # QIcon, QPixmap,
from PyQt5.QtWidgets import *
from PyQt5 import QtCore, QtGui
from database_hadlers.database_handlers_main import get_db_connection, \
    insert_product_data, insert_or_update_products, parse_db_all_products, update_products_dec, \
    insert_prod_menu, select_menu_data, insert_prod_bread, select_bread_data

get_db_connection(path_to_db_file='database/prod_database.db')
locale.setlocale(locale.LC_ALL, "ru_RU")
date = datetime.today().strftime("%d.%m.%Y")
first_month_date = datetime.now().replace(day=1).strftime('%d.%m.%Y')

HEADER_LABELS = ['Найменування', 'од. виміру', 'кількість']
HEADER_LABELS_FOR_MENU = ['Найменування', 'од.виміру', f'наявність на {date}', 'вимагається']
COLUMNS_BREAD = ['Дата', 'Витрачено \nборошна', 'Отримано \nхліба', 'Вихід \nплановий \n(%)', 'Вихід \nфактичний \n(%)', 'Олія\nза нормою\nв кг', 'Олія\nза нормою\nв %', 'Олія\nфактично\nв кг', 'Олія\nфактично\nв %', 'Сіль\nза нормою\nв кг', 'Сіль\nза нормою\nв %', 'Сіль\nфактично\nв кг', 'Сіль\nфактично\nв %', 'Дріжджі\nза нормою\nв кг', 'Дріжджі\nза нормою\nв %', 'Дріжджі\nфактично\nв кг', 'Дріжджі\nфактично\nв %']
COLUMNS_BREAD_ACT = ['Найменування \nматеріальних \nзасобів', 'Одиниця \nвиімру', 'Витрачено \nсировини', 'ціна \nза од.', 'Отримано \nпродукції', 'ціна \nза од.']
ROWS_BREAD_ACT = ['Борошно пшеничне І гат', 'Дріжджі сухі', 'Олія', 'Сіль', 'Хліб пшеничний з борошна І гат.', 'ВСЬОГО:']
COLUMNS_BREAD_ZVIT = ['Дата', 'Витрачено \nборошна', 'Отримано \nхліба', 'Олія\nв кг', 'Сіль\nв кг', 'Дріжджі\nв кг', 'Сума']


class LossProfitTab(QWidget):
    def __init__(self, parent=None):
        self.parent = parent
        super(LossProfitTab, self).__init__()
        self.label_date = QLabel(self)
        self.label_date.setText('Введіть дату операції:')
        self.input_date = QDateEdit(self)
        self.input_date.setCalendarPopup(True)
        self.input_date.setDate(datetime.today())
        self.label_type_oper = QLabel(self)
        self.label_type_oper.setText('Введіть тип операції:')
        self.input_type_oper = QComboBox(self)
        self.input_type_oper.addItems(['Прибуток', 'Убуток'])
        self.label_source_name = QLabel(self)
        self.label_source_name.setText('Назва постачальника (назва в/ч)')
        self.input_source_name = QLineEdit()
        self.label_source_number = QLabel(self)
        self.label_source_number.setText('Номер постачальника (номер в/ч)')
        self.input_source_number = QLineEdit()
        self.label_destination_name = QLabel(self)
        self.label_destination_name.setText('Назва отримувача (назва в/ч)')
        self.input_destination_name = QLineEdit()
        self.label_destination_number = QLabel(self)
        self.label_destination_number.setText('Номер отримувача (номер в/ч)')
        self.input_destination_number = QLineEdit()
        self.label_product_name = QLabel(self)
        self.label_product_name.setText('Назва продукції')
        self.input_product_name = QLineEdit()
        self.label_unit = QLabel(self)
        self.label_unit.setText('  Одиниця виміру  ')
        self.input_unit = QComboBox(self)
        self.input_unit.addItems(['кг', 'шт', 'к-т', 'л', 'др'])
        self.label_product_quantity = QLabel(self)
        self.label_product_quantity.setText('Кількість')
        self.input_product_quantity = QLineEdit()
        self.label_product_price = QLabel(self)
        self.label_product_price.setText('Ціна')
        self.input_product_price = QLineEdit()
        self.label_product_total = QLabel(self)
        self.label_product_total.setText('Сума')
        self.input_product_total = QLineEdit()
        self.label_manufacturer = QLabel(self)
        self.label_manufacturer.setText('Виробник')
        self.input_manufacturer = QLineEdit()
        self.label_production_date = QLabel(self)
        self.label_production_date.setText('    Дата виробництва    ')
        self.input_production_date = QDateEdit(self)
        self.input_production_date.setCalendarPopup(True)
        self.input_production_date.setDate(datetime.today())
        self.label_expiration_date = QLabel(self)
        self.label_expiration_date.setText('Кінцева дата споживання')
        self.input_expiration_date = QDateEdit(self)
        self.input_expiration_date.setCalendarPopup(True)
        self.input_expiration_date.setDate(datetime.today())
        self.label_number_document = QLabel(self)
        self.label_number_document.setText('Номер первинного документу (накладна / акт)')
        self.input_number_document = QLineEdit()
        self.label_date_document = QLabel(self)
        self.label_date_document.setText(' Дата первинного документу ')
        self.input_date_document = QDateEdit(self)
        self.input_date_document.setCalendarPopup(True)
        self.input_date_document.setDate(datetime.today())
        self.label_number_directive = QLabel(self)
        self.label_number_directive.setText('Номер документу(розпорядження / договір)')
        self.input_number_directive = QLineEdit()
        self.label_date_directive = QLabel(self)
        self.label_date_directive.setText('     Дата документу     ')
        self.input_date_directive = QDateEdit(self)
        self.input_date_directive.setCalendarPopup(True)
        self.input_date_directive.setDate(datetime.today())
        # create buttons
        self.save_to_db = QPushButton('   Зберегти у Базу Даних')
        self.save_to_db.setIcon(QtGui.QIcon('icons/database.png'))
        self.save_to_db.clicked.connect(self.push_to_database)

        # layout box
        main_layout = QVBoxLayout(self)
        # 1 row
        input_form_layout = QHBoxLayout(self)
        input_form_layout.addWidget(self.label_date)
        input_form_layout.addWidget(self.input_date)
        input_form_layout.addSpacing(200)
        input_form_layout.addWidget(self.label_type_oper)
        input_form_layout.addWidget(self.input_type_oper)
        # 2 row
        input_source_dest_layout = QHBoxLayout(self)
        input_source_name_layout = QVBoxLayout(self)
        input_source_name_layout.addWidget(self.label_source_name)
        input_source_name_layout.addWidget(self.input_source_name)
        input_source_num_layout = QVBoxLayout(self)
        input_source_num_layout.addWidget(self.label_source_number)
        input_source_num_layout.addWidget(self.input_source_number)
        input_dest_name_layout = QVBoxLayout(self)
        input_dest_name_layout.addWidget(self.label_destination_name)
        input_dest_name_layout.addWidget(self.input_destination_name)
        input_dest_num_layout = QVBoxLayout(self)
        input_dest_num_layout.addWidget(self.label_destination_number)
        input_dest_num_layout.addWidget(self.input_destination_number)
        input_source_dest_layout.addLayout(input_source_name_layout)
        input_source_dest_layout.addLayout(input_source_num_layout)
        input_source_dest_layout.addLayout(input_dest_name_layout)
        input_source_dest_layout.addLayout(input_dest_num_layout)
        # 3 row
        input_product = QHBoxLayout(self)
        input_name_product = QVBoxLayout(self)
        input_name_product.addWidget(self.label_product_name)
        input_name_product.addWidget(self.input_product_name)
        input_product_unit_pr_ttl = QHBoxLayout(self)
        input_unit_layout = QVBoxLayout(self)
        input_unit_layout.addWidget(self.label_unit)
        input_unit_layout.addWidget(self.input_unit)
        input_quantity_layout = QVBoxLayout(self)
        input_quantity_layout.addWidget(self.label_product_quantity)
        input_quantity_layout.addWidget(self.input_product_quantity)
        input_price_layout = QVBoxLayout(self)
        input_price_layout.addWidget(self.label_product_price)
        input_price_layout.addWidget(self.input_product_price)
        input_total_layout = QVBoxLayout(self)
        input_total_layout.addWidget(self.label_product_total)
        input_total_layout.addWidget(self.input_product_total)
        input_product_unit_pr_ttl.addLayout(input_unit_layout)
        input_product_unit_pr_ttl.addLayout(input_quantity_layout)
        input_product_unit_pr_ttl.addLayout(input_price_layout)
        input_product_unit_pr_ttl.addLayout(input_total_layout)
        input_product.addLayout(input_name_product)
        input_product.addLayout(input_product_unit_pr_ttl)
        # 4 row
        input_mf_date_layout = QHBoxLayout(self)
        input_mf_layout = QVBoxLayout(self)
        input_mf_layout.addWidget(self.label_manufacturer)
        input_mf_layout.addWidget(self.input_manufacturer)
        input_pr_date_layout = QVBoxLayout(self)
        input_pr_date_layout.addWidget(self.label_production_date)
        input_pr_date_layout.addWidget(self.input_production_date)
        input_exp_date_layout = QVBoxLayout(self)
        input_exp_date_layout.addWidget(self.label_expiration_date)
        input_exp_date_layout.addWidget(self.input_expiration_date)
        input_mf_date_layout.addLayout(input_mf_layout)
        input_mf_date_layout.addLayout(input_pr_date_layout)
        input_mf_date_layout.addLayout(input_exp_date_layout)
        # 5 row
        input_documents = QHBoxLayout(self)
        input_num_doc_layout = QVBoxLayout(self)
        input_num_doc_layout.addWidget(self.label_number_document)
        input_num_doc_layout.addWidget(self.input_number_document)
        input_date_doc_layout = QVBoxLayout(self)
        input_date_doc_layout.addWidget(self.label_date_document)
        input_date_doc_layout.addWidget(self.input_date_document)
        input_num_dir_layout = QVBoxLayout(self)
        input_num_dir_layout.addWidget(self.label_number_directive)
        input_num_dir_layout.addWidget(self.input_number_directive)
        input_date_dir_layout = QVBoxLayout(self)
        input_date_dir_layout.addWidget(self.label_date_directive)
        input_date_dir_layout.addWidget(self.input_date_directive)
        input_documents.addLayout(input_num_doc_layout)
        input_documents.addLayout(input_date_doc_layout)
        input_documents.addLayout(input_num_dir_layout)
        input_documents.addLayout(input_date_dir_layout)
        # button
        button_layout = QHBoxLayout(self)
        button_layout.addWidget(self.save_to_db)

        main_layout.addLayout(input_form_layout)
        main_layout.addSpacing(15)
        main_layout.addLayout(input_source_dest_layout)
        main_layout.addSpacing(20)
        main_layout.addLayout(input_product)
        main_layout.addSpacing(15)
        main_layout.addLayout(input_mf_date_layout)
        main_layout.addSpacing(15)
        main_layout.addLayout(input_documents)
        main_layout.addStretch()
        main_layout.addLayout(button_layout)

    def push_to_database(self):
        data_tuple = (self.input_source_name.text(),
                      self.input_source_number.text(),
                      self.input_destination_name.text(),
                      self.input_destination_number.text(),
                      self.input_date.text(),
                      self.input_product_name.text(),
                      self.input_unit.currentText(),
                      self.input_product_quantity.text(),
                      self.input_product_price.text(),
                      self.input_product_total.text(),
                      self.input_type_oper.currentText(),
                      self.input_manufacturer.text(),
                      self.input_production_date.text(),
                      self.input_expiration_date.text(),
                      self.input_number_document.text(),
                      self.input_date_document.text(),
                      self.input_number_directive.text(),
                      self.input_date_directive.text())
        insert_product_data(data_tuple)
        data_tuple_2 = (self.input_product_name.text(),
                        self.input_unit.currentText(),
                        self.input_product_quantity.text(),
                        self.input_type_oper.currentText())
        insert_or_update_products(data_tuple_2)


class Storage(QWidget):
    def __init__(self, parent=None):
        super(Storage, self).__init__()
        self.parent = parent
        # create table widget
        self.table_widget = QTableWidget(0, 3) # rows, columns
        self.table_widget.setHorizontalHeaderLabels(HEADER_LABELS) # headers of columns on table
        self.table_widget.horizontalHeader().setDefaultSectionSize(200)
        self.table_widget.setColumnWidth(0, 350)
        # create button
        self.push_button = QPushButton('   Сформувати таблицю')
        self.push_button.setIcon(QtGui.QIcon('icons/computer.png'))
        self.push_button.clicked.connect(self.show_table_func)
        self.form_excel = QPushButton('   Формувати у Excel')
        self.form_excel.setIcon(QtGui.QIcon('icons/excel.png'))
        self.form_excel.clicked.connect(self.export_to_excel)
        # create dialog-window for save file
        self.dialog = QFileDialog(self)
        # layout box
        main_layout = QVBoxLayout(self)
        main_layout.addWidget(self.table_widget)
        main_layout.addWidget(self.push_button)
        main_layout.addWidget(self.form_excel)

    def show_table_func(self):
        '''
        function for create and show data from 'main_file' table
        '''
        self.all_producrs = parse_db_all_products()
        self.table_widget.setRowCount(len(self.all_producrs))
        row_table = 0
        for row in self.all_producrs:
            column_table = 0
            for column in row[1:]:
                self.table_widget.setItem(row_table, column_table, QTableWidgetItem(str(column)))
                column_table += 1
            row_table += 1

    def export_to_excel(self):
        column_headers = []
        row_count = self.table_widget.model().rowCount()
        for j in range(self.table_widget.model().columnCount()):
            column_headers.append(self.table_widget.horizontalHeaderItem(j).text())
            df = pd.DataFrame(columns=column_headers)
        for row in range(row_count):
            for col in range(self.table_widget.columnCount()):
                try:
                    temp = self.table_widget.item(row, col).text()
                except:
                    temp = 0
                try:
                    temp = float(temp)
                    temp = locale.str(temp)
                except:
                    pass
                df.at[row, column_headers[col]] = temp

        # activate dialog-window for save file
        result = self.dialog.getSaveFileName(self.table_widget, 'Зберегти файл', 'C:/', 'Excel files (*.xlsx)')
        # try-except block for saving file
        try:
            df.to_excel(result[0])
        except:
            pass


class Menu(QWidget):
    def __init__(self, parent=None):
        super(Menu, self).__init__()
        self.parent = parent
        # create date widget's
        self.label_op_date = QLabel(self)
        self.label_op_date.setText('Введіть дату операції:')
        self.input_op_date = QDateEdit(self)
        self.input_op_date.setCalendarPopup(True)
        self.input_op_date.setDate(datetime.today())
        self.label_menu_date = QLabel(self)
        self.label_menu_date.setText('Введіть дату на яку здійснюється операція:')
        self.input_menu_date = QDateEdit(self)
        self.input_menu_date.setCalendarPopup(True)
        self.input_menu_date.setDate(datetime.today())
        # create table widget
        self.table_widget = QTableWidget(0, 4) # rows, columns
        self.table_widget.setHorizontalHeaderLabels(HEADER_LABELS_FOR_MENU) # headers of columns on table
        self.table_widget.horizontalHeader().setDefaultSectionSize(200)
        self.table_widget.setColumnWidth(0, 350)
        # create button
        self.push_button = QPushButton('   Сформувати таблицю')
        self.push_button.setIcon(QtGui.QIcon('icons/computer.png'))
        self.push_button.clicked.connect(self.show_table_func)
        self.save_to_db = QPushButton('   Зберегти у Базу Даних')
        self.save_to_db.setIcon(QtGui.QIcon('icons/database.png'))
        self.save_to_db.clicked.connect(self.push_to_database)
        self.form_excel = QPushButton('   Формувати у Excel')
        self.form_excel.setIcon(QtGui.QIcon('icons/excel.png'))
        self.form_excel.clicked.connect(self.export_to_excel)
        # create dialog-window for save file
        self.dialog = QFileDialog(self)
        # layout box
        main_layout = QVBoxLayout(self)
        # 1 row
        input_form_layout = QHBoxLayout(self)
        input_form_layout.addWidget(self.label_op_date)
        input_form_layout.addWidget(self.input_op_date)
        input_form_layout.addSpacing(200)
        input_form_layout.addWidget(self.label_menu_date)
        input_form_layout.addWidget(self.input_menu_date)
        # 2 row
        table_layout = QHBoxLayout(self)
        table_layout.addWidget(self.table_widget)
        # 3 row
        button_layout = QVBoxLayout(self)
        button_layout.addWidget(self.push_button)
        button_layout.addWidget(self.save_to_db)
        button_layout.addWidget(self.form_excel)
        # main_layout
        main_layout.addLayout(input_form_layout)
        main_layout.addWidget(self.table_widget)
        main_layout.addLayout(button_layout)

    def show_table_func(self):
        '''
        function for create and show data from 'main_file' table
        '''
        self.all_producrs = parse_db_all_products()
        self.table_widget.setRowCount(len(self.all_producrs))
        row_table = 0
        for row in self.all_producrs:
            column_table = 0
            for column in row[1:]:
                self.table_widget.setItem(row_table, column_table, QTableWidgetItem(str(column)))
                column_table += 1
            row_table += 1

    def push_to_database(self):
        date_op = (self.input_op_date.text(),)
        date_menu = (self.input_menu_date.text(),)
        data_for_all_prod = []
        data_for_menu_prod = []
        for row in range(self.table_widget.model().rowCount()):
            tuple_for_all_prod = []
            for column in range(self.table_widget.model().columnCount()):
                if self.table_widget.item(row, column) is not None:
                    item = self.table_widget.item(row, column).text()
                else:
                    item = 0
                if item == '':
                    item = 0
                tuple_for_all_prod.append(item)
            tuple_for_all_prod = tuple(tuple_for_all_prod)
            if tuple_for_all_prod[3] != 0:
                data_for_all_prod.append(tuple_for_all_prod)
            tuple_for_all_prod = date_op + date_menu + tuple_for_all_prod
            data_for_menu_prod.append(tuple_for_all_prod)
        for i in data_for_all_prod:
            update_products_dec(list(i))
        insert_prod_menu(data_for_menu_prod)

    def export_to_excel(self):
        column_headers = []
        row_count = self.table_widget.model().rowCount()
        for j in range(self.table_widget.model().columnCount()):
            column_headers.append(self.table_widget.horizontalHeaderItem(j).text())
            df = pd.DataFrame(columns=column_headers)
        for row in range(row_count):
            for col in range(self.table_widget.columnCount()):
                try:
                    temp = self.table_widget.item(row, col).text()
                except:
                    temp = 0
                try:
                    temp = float(temp)
                    temp = locale.str(temp)
                except:
                    pass
                df.at[row, column_headers[col]] = temp

        # activate dialog-window for save file
        result = self.dialog.getSaveFileName(self.table_widget, 'Зберегти файл', 'C:/', 'Excel files (*.xlsx)')
        # try-except block for saving file
        try:
            df.to_excel(result[0])
        except:
            pass


class MenuReport(QWidget):
    def __init__(self, parent=None):
        self.parent = parent
        super(MenuReport, self).__init__()
        # create date widget's
        self.label_start_date = QLabel(self)
        self.label_start_date.setText('Введіть початкову дату звіту:')
        self.input_start_date = QDateEdit(self)
        self.input_start_date.setCalendarPopup(True)
        self.input_start_date.setDate(datetime.now().replace(day=1))

        self.label_end_date = QLabel(self)
        self.label_end_date.setText('Введіть завершаючу дату звіту:')
        self.input_end_date = QDateEdit(self)
        self.input_end_date.setCalendarPopup(True)
        self.input_end_date.setDate(datetime.today())
        # create table widget
        self.table_widget = QTableWidget()
        # create button
        self.push_button = QPushButton('   Сформувати таблицю')
        self.push_button.setIcon(QtGui.QIcon('icons/computer.png'))
        self.push_button.clicked.connect(self.show_table_func)
        self.form_excel = QPushButton('   Формувати у Excel')
        self.form_excel.setIcon(QtGui.QIcon('icons/excel.png'))
        self.form_excel.clicked.connect(self.export_to_excel)
        # create dialog-window for save file
        self.dialog = QFileDialog(self)
        # layout box
        main_layout = QVBoxLayout(self)
        # 1 row
        input_form_layout = QHBoxLayout(self)
        input_form_layout.addWidget(self.label_start_date)
        input_form_layout.addWidget(self.input_start_date)
        input_form_layout.addSpacing(200)
        input_form_layout.addWidget(self.label_end_date)
        input_form_layout.addWidget(self.input_end_date)
        # 2 row
        table_layout = QHBoxLayout(self)
        table_layout.addWidget(self.table_widget)
        # 3 row
        button_layout = QVBoxLayout(self)
        button_layout.addWidget(self.push_button)
        button_layout.addWidget(self.form_excel)
        # main_layout
        main_layout.addLayout(input_form_layout)
        main_layout.addWidget(self.table_widget)
        main_layout.addLayout(button_layout)

    def parse_column_names(self):
        data_query = select_menu_data((self.input_start_date.text(), self.input_end_date.text()))
        data_for_column_names = []
        for i in data_query:
            name = i[3]
            if name not in data_for_column_names:
                data_for_column_names.append(name)
        return data_for_column_names

    def parse_menu_data(self):
        data_query = select_menu_data((self.input_start_date.text(), self.input_end_date.text()))
        start_date = data_query[0][7]
        menu_data = []
        data_for_date = []
        for i in data_query:
            data = (i[2], i[3], i[6])
            if start_date == i[7]:
                data_for_date.append(data)
            else:
                inner_list = copy.copy(data_for_date)
                menu_data.append(inner_list)
                data_for_date.clear()
                data_for_date.append(data)
                start_date = i[7]
        menu_data.append(data_for_date)
        return menu_data

    def show_table_func(self):
        column_names = self.parse_column_names()
        menu_data = self.parse_menu_data()
        self.table_widget.setColumnCount(len(column_names)+1)
        self.table_widget.setHorizontalHeaderLabels(['Дата']+column_names)
        self.table_widget.setRowCount(len(menu_data)+1)
        row_table = 0
        for i in menu_data:
            column_table = 0
            date_for_row = i[0][0]
            self.table_widget.setItem(row_table, 0, QTableWidgetItem(str(date_for_row)))
            for j in i:
                self.table_widget.setItem(row_table, column_names.index(j[1])+1, QTableWidgetItem(str(j[2])))
                column_table += 1
            row_table += 1
        sum_columns = ['Всього:']

        for column in range(1, self.table_widget.columnCount()):
            sum_column = []
            for row in range(0, self.table_widget.rowCount()):
                if self.table_widget.item(row, column) is not None:
                    item = float(self.table_widget.item(row, column).text())
                else:
                    item = 0
                sum_column.append(item)
            sum_columns.append(round(sum(sum_column), 3))
        column = 0
        for i in sum_columns:
            self.table_widget.setItem(self.table_widget.rowCount()-1, column, QTableWidgetItem(str(i)))
            column += 1

    def export_to_excel(self):
        column_headers = []
        row_count = self.table_widget.model().rowCount()
        for j in range(self.table_widget.model().columnCount()):
            column_headers.append(self.table_widget.horizontalHeaderItem(j).text())
            df = pd.DataFrame(columns=column_headers)
        for row in range(row_count):
            for col in range(self.table_widget.columnCount()):
                try:
                    temp = self.table_widget.item(row, col).text()
                except:
                    temp = 0
                try:
                    temp = float(temp)
                    temp = locale.str(temp)
                except:
                    pass
                df.at[row, column_headers[col]] = temp

        # activate dialog-window for save file
        result = self.dialog.getSaveFileName(self.table_widget, 'Зберегти файл', 'C:/', 'Excel files (*.xlsx)')
        # try-except block for saving file
        try:
            df.to_excel(result[0])
        except:
            pass


class Bread(QWidget):
    def __init__(self, parent=None):
        super(Bread, self).__init__()
        self.parent = parent
        # create input fields
        self.label_date = QLabel(self)
        self.label_date.setText('Введіть дату операції:')
        self.input_date = QDateEdit(self)
        self.input_date.setCalendarPopup(True)
        self.input_date.setDate(datetime.today())
        self.label_bread = QLabel(self)
        self.label_bread.setText('Введіть кількість продукції:')
        self.input_bread = QLineEdit('0')
        self.label_coeff = QLabel(self)
        self.label_coeff.setText('Введіть коефіцієнт виходу:')
        self.input_coeff = QLineEdit('136.1')
        # create tables
        self.table_widget = QTableWidget(0, 17)  # +1
        self.table_widget.setHorizontalHeaderLabels(COLUMNS_BREAD)
        self.table_widget.horizontalHeader().setDefaultSectionSize(80)
        self.table_widget_2 = QTableWidget(0, 6)  # +1
        self.table_widget_2.setHorizontalHeaderLabels(COLUMNS_BREAD_ACT)
        self.table_widget_2.horizontalHeader().setDefaultSectionSize(160)
        self.table_widget_2.setColumnWidth(0, 300)
        # create buttons
        self.form_table_button = QPushButton('   Сформувати таблицю')
        self.form_table_button.setIcon(QtGui.QIcon('icons/computer.png'))
        self.form_table_button.clicked.connect(self.show_table_func)
        self.calculate_button = QPushButton('   Провести розрахунок')
        self.calculate_button.setIcon(QtGui.QIcon('icons/calculate.png'))
        self.calculate_button.clicked.connect(self.calculate_result)
        self.save_to_db_button = QPushButton('   Зберегти у Базу Даних')
        self.save_to_db_button.setIcon(QtGui.QIcon('icons/database.png'))
        self.save_to_db_button.clicked.connect(self.push_to_database)
        self.excel_button = QPushButton('   Формувати у Excel')
        self.excel_button.setIcon(QtGui.QIcon('icons/excel.png'))
        self.excel_button.clicked.connect(self.export_to_excel)
        # create dialog-window for save file
        self.dialog = QFileDialog(self)

        main_layout = QVBoxLayout(self)
        # 1 row
        input_form_layout = QHBoxLayout(self)
        input_form_layout.addWidget(self.label_date)
        input_form_layout.addWidget(self.input_date)
        input_form_layout.addWidget(self.label_bread)
        input_form_layout.addWidget(self.input_bread)
        input_form_layout.addWidget(self.label_coeff)
        input_form_layout.addWidget(self.input_coeff)
        button_layout = QHBoxLayout(self)
        button_layout.addWidget(self.form_table_button)
        button_layout.addWidget(self.calculate_button)
        button_layout.addWidget(self.save_to_db_button)
        button_layout.addWidget(self.excel_button)
        main_layout.addLayout(input_form_layout)
        main_layout.addWidget(self.table_widget)
        main_layout.addWidget(self.table_widget_2)
        main_layout.addLayout(button_layout)
        # add dialog-window for save file
        main_layout.addWidget(self.dialog)

    def show_table_func(self):
        self.table_widget.setRowCount(1)
        self.table_widget_2.setRowCount(6)
        date = self.input_date.text()
        bread = float(self.input_bread.text())
        out_p = float(self.input_coeff.text()) / 100
        oil_p = 0.141
        salt_p = 1.8
        yeast_p = 0.4
        wheat = round(bread / out_p, 3)
        oil = round((wheat * oil_p) / 100, 3)
        salt = round((wheat * salt_p) / 100, 3)
        yeast = round((wheat * yeast_p) / 100, 3)

        values_list = [date, wheat, bread, out_p, out_p, oil, oil_p, oil, oil_p, salt, salt_p, salt, salt_p, yeast, yeast_p, yeast, yeast_p]

        names_numb = 0
        for column in range(17):
            self.table_widget.setItem(0, column, QTableWidgetItem(str(values_list[names_numb])))
            names_numb = names_numb+1

        row = 0
        for row_item in ROWS_BREAD_ACT:
            self.table_widget_2.setItem(row, 0, QTableWidgetItem(str(row_item)))
            row = row+1

        for row in range(5):
            self.table_widget_2.setItem(row, 1, QTableWidgetItem('кг'))

        self.table_widget_2.setItem(0, 2, QTableWidgetItem(str(wheat)))
        self.table_widget_2.setItem(1, 2, QTableWidgetItem(str(yeast)))
        self.table_widget_2.setItem(2, 2, QTableWidgetItem(str(oil)))
        self.table_widget_2.setItem(3, 2, QTableWidgetItem(str(salt)))
        self.table_widget_2.setItem(4, 4, QTableWidgetItem(str(bread)))

    def calculate_result(self):
        sum_ingredients = []
        for row in range(5):
            try:
                item = round(float(self.table_widget_2.item(row, 2).text()), 3)
            except:
                item = 0.0
            try:
                item_price = round(float(self.table_widget_2.item(row, 3).text()), 2)
            except:
                item_price = 0.0
            item_sum = round(item * item_price, 2)
            sum_ingredients.append(item_sum)
        sum_ingredients = round(sum(sum_ingredients), 2)
        self.table_widget_2.setItem(5, 3, QTableWidgetItem(str(sum_ingredients)))
        self.table_widget_2.setItem(5, 5, QTableWidgetItem(str(sum_ingredients)))
        try:
            bread_price = round(sum_ingredients / round(float(self.table_widget_2.item(4, 4).text()), 3), 2)
        except:
            bread_price = 0
        self.table_widget_2.setItem(4, 5, QTableWidgetItem(str(bread_price)))

    def push_to_database(self):
        date_op = [self.input_date.text()]
        bread = [self.input_bread.text()]
        sum_bread = [self.table_widget_2.item(5, 5).text()]
        data_for_all_prod = []
        for row in range(self.table_widget_2.model().rowCount()):
            tuple_for_all_prod = []
            for column in range(self.table_widget_2.model().columnCount()):
                if self.table_widget_2.item(row, column) is not None:
                    item = self.table_widget_2.item(row, column).text()
                else:
                    item = 0
                if item == '':
                    item = 0
                tuple_for_all_prod.append(item)
            tuple_for_all_prod = date_op + tuple_for_all_prod
            data_for_all_prod.append(tuple_for_all_prod)
        ingredients = []
        for i in data_for_all_prod[:4]:
            i[0], i[1] = i[1], i[0]
            update_products_dec(i)
            i[3] = abs(float(i[3]))
            ingredients.append(i[3])
        bread = date_op + bread + ingredients + sum_bread
        insert_prod_bread(bread)
        bread_for_all_products = [self.table_widget_2.item(4, 0).text()] + [self.table_widget_2.item(4, 1).text()] + [self.table_widget_2.item(4, 4).text()] + ['Прибуток']
        insert_or_update_products(bread_for_all_products)

    def export_to_excel(self):
        column_headers = []
        row_count = self.table_widget_2.model().rowCount()
        for j in range(self.table_widget_2.model().columnCount()):
            column_headers.append(self.table_widget_2.horizontalHeaderItem(j).text())
            df = pd.DataFrame(columns=column_headers)
        for row in range(row_count):
            for col in range(self.table_widget_2.columnCount()):
                try:
                    temp = self.table_widget_2.item(row, col).text()
                except:
                    temp = 0
                try:
                    temp = float(temp)
                    temp = locale.str(temp)
                except:
                    pass
                df.at[row, column_headers[col]] = temp

        # activate dialog-window for save file
        result = self.dialog.getSaveFileName(self.table_widget_2, 'Зберегти файл', 'C:/', 'Excel files (*.xlsx)')
        # try-except block for saving file
        try:
            df.to_excel(result[0])
        except:
            pass


class BreadZvit(QWidget):
    def __init__(self, parent=None):
        super(BreadZvit, self).__init__()
        self.parent = parent
        # create table
        self.table_widget = QTableWidget(0, 7)
        self.table_widget.setHorizontalHeaderLabels(COLUMNS_BREAD_ZVIT)
        # create input date labels
        self.label_date_1 = QLabel(self)
        self.label_date_1.setText('Введіть початкову дату операції:')
        self.input_date_1 = QDateEdit(self)
        self.input_date_1.setCalendarPopup(True)
        self.input_date_1.setDate(datetime.now().replace(day=1))
        self.label_date_2 = QLabel(self)
        self.label_date_2.setText('Введіть кіневу дату операції:')
        self.input_date_2 = QDateEdit(self)
        self.input_date_2.setCalendarPopup(True)
        self.input_date_2.setDate(datetime.today())
        # create buttons
        self.form_table_button = QPushButton('   Сформувати таблицю')
        self.form_table_button.setIcon(QtGui.QIcon('icons/computer.png'))
        self.form_table_button.clicked.connect(self.show_table_func)
        self.excel_button = QPushButton('   Формувати у Excel')
        self.excel_button.setIcon(QtGui.QIcon('icons/excel.png'))
        self.excel_button.clicked.connect(self.export_to_excel)
        # create dialog-window for save file
        self.dialog = QFileDialog(self)

        main_layout = QVBoxLayout(self)
        input_form_layout = QHBoxLayout(self)
        input_form_layout.addWidget(self.label_date_1)
        input_form_layout.addWidget(self.input_date_1)
        input_form_layout.addWidget(self.label_date_2)
        input_form_layout.addWidget(self.input_date_2)
        button_layout = QHBoxLayout(self)
        button_layout.addWidget(self.form_table_button)
        button_layout.addWidget(self.excel_button)
        main_layout.addLayout(input_form_layout)
        main_layout.addWidget(self.table_widget)
        main_layout.addLayout(button_layout)
        # add dialog-window for save file
        main_layout.addWidget(self.dialog)

    def show_table_func(self):
        dates = self.input_date_1.text(), self.input_date_2.text()
        data = select_bread_data(dates)
        self.table_widget.setRowCount(len(data) + 1)
        row_table = 0
        for i in data:
            column_table = 0
            for j in i[1:]:
                self.table_widget.setItem(row_table, column_table, QTableWidgetItem(str(j)))
                column_table += 1
            row_table += 1
        sum_columns = ['Всього:']
        for column in range(1, self.table_widget.columnCount()):
            sum_column = []
            for row in range(0, self.table_widget.rowCount()):
                if self.table_widget.item(row, column) is not None:
                    item = float(self.table_widget.item(row, column).text())
                else:
                    item = 0
                sum_column.append(item)
            sum_columns.append(round(sum(sum_column), 3))
        column = 0
        for i in sum_columns:
            self.table_widget.setItem(self.table_widget.rowCount()-1, column, QTableWidgetItem(str(i)))
            column += 1

    def export_to_excel(self):
        column_headers = []
        row_count = self.table_widget.model().rowCount()
        for j in range(self.table_widget.model().columnCount()):
            column_headers.append(self.table_widget.horizontalHeaderItem(j).text())
            df = pd.DataFrame(columns=column_headers)
        for row in range(row_count):
            for col in range(self.table_widget.columnCount()):
                try:
                    temp = self.table_widget.item(row, col).text()
                except:
                    temp = 0
                try:
                    temp = float(temp)
                    temp = locale.str(temp)
                except:
                    pass
                df.at[row, column_headers[col]] = temp
        # activate dialog-window for save file
        result = self.dialog.getSaveFileName(self.table_widget, 'Зберегти файл', 'C:/', 'Excel files (*.xlsx)')
        # try-except block for saving file
        try:
            df.to_excel(result[0])
        except:
            pass


class ReadWriteExcel(QWidget):
    def __init__(self, parent=None):
        self.parent = parent
        super(ReadWriteExcel, self).__init__()
        self.make_file = QPushButton('   Почати заповнення "Залишки підрозділи"')
        self.make_file.setIcon(QtGui.QIcon('img/process.png'))
        self.make_file.clicked.connect(self.scrap_write)
        self.make_file_2 = QPushButton('   Почати заповнення "Залишки ООДК"')
        self.make_file_2.setIcon(QtGui.QIcon('img/process.png'))
        self.make_file_2.clicked.connect(self.scrap_write_2)
        main_v_box = QVBoxLayout(self)
        main_v_box.addWidget(self.make_file)
        main_v_box.addWidget(self.make_file_2)

    def get_dir_path(self):
        self.files_directory_choice = QFileDialog.getExistingDirectory(self, 'Оберіть папку з вихідними файлами')
        return self.files_directory_choice

    def get_file_path(self):
        self.file_main = QFileDialog.getOpenFileName(self, 'Оберіть базовий файл')
        return self.file_main[0]

    def scrap_write(self):
        f_dir = self.get_dir_path()
        m_file = self.get_file_path()
        files = os.listdir(f_dir)
        val_massive = []
        for file in files:
            read_data = openpyxl.load_workbook(f_dir+'/'+file)
            sheet_1 = read_data.sheetnames[0]
            sheet = read_data[sheet_1]
            columns = [2, 3, 4, 5, 10, 12]
            for coll in columns:
                mark = sheet.cell(column=coll, row=6).value
                if mark == None:
                    break
                val_list = []
                count = 0
                for row in range(6, 119):
                    val = sheet.cell(column=coll, row=row).value
                    if val == None:
                        val = 0
                    if type(val) == float or type(val) == int:
                        if count > 22:
                            if val == 0:
                                val = '0,000'
                            else:
                                val = round((val / 1000), 3)
                                val = locale.str(val)
                    count += 1
                    val_list.append(val)
                val_massive.append(val_list)
        open_file = openpyxl.load_workbook(m_file)
        sheet_main_1 = open_file.sheetnames[0]
        sheet_main = open_file[sheet_main_1]
        columns_main = [x for x in range(4, (len(val_massive)+2)*2) if x % 2 == 0]
        col_index = 0
        for val_l in val_massive:
            column_numb = columns_main[col_index]
            row = 6
            for val in val_l:
                sheet_main.cell(column=column_numb, row=row).value = val
                row += 1
            col_index += 1
        new_filename = m_file.split('.')
        new_filename.insert(1, date)
        new_filename[2] = '.xlsx'
        new_filename = ' '.join(new_filename)
        open_file.save(new_filename)

    # for button #2 functions
    def get_src_file_path(self):
        self.files_directory_choice = QFileDialog.getOpenFileName(self, 'Оберіть вихідний файл')
        return self.files_directory_choice[0]

    def scrap_write_2(self):
        src_file = self.get_src_file_path()
        base_file = self.get_file_path()
        read_data_src_file = openpyxl.load_workbook(src_file)
        sheet_1_src_file = read_data_src_file.sheetnames[0]
        sheet_src_file = read_data_src_file[sheet_1_src_file]
        dict_source_1 = {}
        dict_source_2 = {}
        for row in range(7, 119):
            key = sheet_src_file.cell(column=1, row=row).value
            val_1 = sheet_src_file.cell(column=32, row=row).value
            val_2 = sheet_src_file.cell(column=33, row=row).value
            dict_source_1[key] = val_1
            dict_source_2[key] = val_2
        read_data_base_file = openpyxl.load_workbook(base_file)
        sheet_1_base_file = read_data_base_file.sheetnames[0]
        sheet_base_file = read_data_base_file[sheet_1_base_file]
        ### 1
        for item in dict_source_1.items():
            name, value = item[0], item[1]
            prev_key = None
            for i in range(6, 117):
                key = sheet_base_file.cell(column=1, row=i).value
                if key == None:
                    key = prev_key
                else:
                    prev_key = key
                if key in ['Капуста', 'Морква', 'Буряк', 'Цибуля', 'Часник', 'Огірки, помідори, баклажани']:
                    key = key + ' ' + sheet_base_file.cell(column=2, row=i).value
                if key.lower() == name.lower():
                    row = sheet_base_file.cell(column=1, row=i).row
                    sheet_base_file.cell(column=3, row=row).value = value
                    break
        ### 2
        for item in dict_source_2.items():
            name, value = item[0], item[1]
            prev_key = None
            for i in range(6, 117):
                key = sheet_base_file.cell(column=1, row=i).value
                if key == None:
                    key = prev_key
                else:
                    prev_key = key
                if key in ['Капуста', 'Морква', 'Буряк', 'Цибуля', 'Часник', 'Огірки, помідори, баклажани']:
                    key = key + ' ' + sheet_base_file.cell(column=2, row=i).value
                if key.lower() == name.lower():
                    row = sheet_base_file.cell(column=1, row=i).row
                    sheet_base_file.cell(column=5, row=row).value = value
                    break
        new_filename = base_file.split('.')
        new_filename.insert(1, date)
        new_filename[2] = '.xlsx'
        new_filename = ' '.join(new_filename)
        read_data_base_file.save(new_filename)


class MainWindow(QMainWindow):
    """
    MAIN window Class
    """
    def __init__(self):
        super().__init__() # initialization widgets and properties Parent class "QDialog"
        # in here we set widgets and set properties
        self.setWindowTitle('eBook')
        self.setWindowIcon(QtGui.QIcon('icons/main.png'))
        # self.setWindowTitle("My App") # title of app
        self.resize(1150, 1000) # set size window
        font = QFont("Times New Roman", 14, 75, True) # set font window
        # add widgets
        self.main_widget = QTabWidget()
        self.setCentralWidget(self.main_widget)
        self.main_widget.addTab(Storage(), "Залишки (Склад)")
        self.main_widget.addTab(LossProfitTab(), "Прихід / Розхід")
        self.main_widget.addTab(Menu(), "Розхід на меню-вимогу")
        self.main_widget.addTab(MenuReport(), "Звіт меню-вимоги")
        self.main_widget.addTab(Bread(), "Хлібопечення")
        self.main_widget.addTab(BreadZvit(), "Хлібопечення звіт")
        self.main_widget.addTab(ReadWriteExcel(), "Парсер Excel")


### Final App Block ###
if __name__ == '__main__':
    app = QApplication(sys.argv)  # create app
    # app.setStyle('Oxygen') # 'Breeze'
    dlgMain = MainWindow() # build object of class "DlgMain" and set here in variable "dlgMain"
    dlgMain.show() # show function
    sys.exit(app.exec_())  # loop app in "sys.exit" func for check logs.

